import ezdxf
from utils.file_utils import save_in_temp_folder
from ezdxf.enums import TextEntityAlignment
import openpyxl
import os

class ListToDxf:
  def __init__(self, file):
    self.full_path = save_in_temp_folder(file, __file__)
    self.base_dir = os.path.dirname(os.path.abspath(__file__))
    self.file_name = os.path.splitext(os.path.basename(self.full_path))[0]
    self.workbook = openpyxl.load_workbook(self.full_path, data_only=True)
    self.drawn_list = self.read_sheets_names()
    self.project_code = self.get_project_code()
    self.target_dxf_path = self.create_DXF_file()
    self.doc_dxf = ezdxf.readfile(self.target_dxf_path)
    self.msp_dxf = self.doc_dxf.modelspace()
    self.block_offset_x = 0
    self.block_offset_y = 0
    
    # Loop para adicionar todas as Listas no DXF
    for index, drawn_list in enumerate(self.drawn_list):
      # Determina a posição dos Textos, Blocos etc.
      self.block_offset_x = 270 * index
      self.block_offset_y = 0

      self.add_material_list(drawn_list.list_material.materials)
      self.add_header_material_list()
      self.add_general_notes(drawn_list.drawing_code_list, drawn_list.list_part.part)
      self.add_part_list(drawn_list.list_part.part)

      # Adiciona um retangulo em volta das informações, para dividir em vários informações
      self.msp_dxf.add_lwpolyline([(self.block_offset_x, 0, 0), (self.block_offset_x, 574, 0), (self.block_offset_x+180, 574, 0), (self.block_offset_x+180, 0, 0), (self.block_offset_x, 0, 0)])
      
      # Adiciona um indentificador do Codigo do Desenho em cima do Retangulo
      self.msp_dxf.add_mtext(
        text= "\\A1;" + drawn_list.drawing_code_list,
        dxfattribs={
          "insert": (self.block_offset_x + 85, 599, 0),  
          "char_height": 20,  
          "width": 0,  
          "style": "Standard",
          "rotation": 0,
          "color": 4,
          "layer" : "COTAS",
          "attachment_point": 5
        }
      )

      # Salva o DXF
      self.doc_dxf.save()
      
  # Função para pegar as Listas do Desenho
  def read_sheets_names(self):
    sheets_names_collection = set()

    for sheet in self.workbook.worksheets[1:]:
      name = sheet.title.split('|')[0].strip()
      sheets_names_collection.add(name)

    sheets_names_array = list(sheets_names_collection)
    sheets_names_array.sort()

    drawing_list_array = []
    for sheet in sheets_names_array:
      drawing_list_array.append(DrawingList(sheet, self.workbook[sheet + ' | PEÇAS'], self.workbook[sheet + ' | MATERIAL']))
    
    return drawing_list_array
  
  # Função para determinar código do Projeto
  def get_project_code(self):
    targetSheet = self.workbook.worksheets[0]
    collumnPieceNumber = targetSheet['E']
    projectCode = None
    highestCount = 0

    countOccurrences = {}
    for cell in collumnPieceNumber[1:]:
      if isinstance(cell.value, str):
        cell_value = cell.value[:7]
        countOccurrences[cell_value] = countOccurrences.get(cell_value, 0) + 1

    for value, count in countOccurrences.items():
      if count > highestCount:
        projectCode = value
        highestCount = count

    return projectCode
    
  # Função para copiar o Template e criar um novo arquivo para Edição
  def create_DXF_file(self):
    output_dir = os.path.join(self.base_dir, "temp")
    os.makedirs(output_dir, exist_ok=True)
    target_dxf_path = os.path.join(output_dir, self.file_name + ".dxf")

    # Caso arquivo exista ele irá remover
    if os.path.isfile(target_dxf_path):
      os.remove(target_dxf_path)

    source_dxf = ezdxf.readfile(os.path.join(self.base_dir, "BLOCO-BRANCO.dxf"))
    source_dxf.saveas(target_dxf_path)
    
    return target_dxf_path

  # Função para adicionar a Lista de Material ao Blocos de Material e adicionar ao DXF
  def add_material_list(self, list_material):
    material_block = self.doc_dxf.blocks.get('BOM')
    if not material_block : return

    self.block_offset_y += 6
    for material in list_material:
      self.block_offset_y += 6
      material_block_insert = self.msp_dxf.add_blockref(material_block.name, insert=(self.block_offset_x, self.block_offset_y, 0))
    
      values = {
        "POS.": material.position,
        "DES": material.description,
        "UNI": material.unit,
        "QTY": material.quantity,
        "MAT": material.composition,
        "WEI": material.total_weight
      }

      material_block_insert.add_auto_attribs(values)
  
  # Função para adicionar o Bloco Descritivo de Material ao DXF
  def add_header_material_list(self):
    header_material_block = self.doc_dxf.blocks.get('INDEX_BOM')
    note_material_block = self.doc_dxf.blocks.get('TOP_BOM')

    self.block_offset_y += 6
    self.msp_dxf.add_blockref(header_material_block.name, insert=(self.block_offset_x, 6, 0))
    self.msp_dxf.add_blockref(note_material_block.name, insert=(self.block_offset_x, self.block_offset_y, 0))
    self.block_offset_y += 11

  # Adiciona os Blocos de Peças ao DXF
  def add_part_list(self, list_part):
    part_block = self.msp_dxf.query('INSERT[name=="REDECAM_STEELWORK"]').first
    header_part_block = self.doc_dxf.blocks.get('INDEX_STEELWORK')
    if not part_block : return

    self.block_offset_y += 6
    self.msp_dxf.add_blockref(header_part_block.name, insert=(self.block_offset_x, self.block_offset_y, 0))

    for part in list_part:
      self.block_offset_y += 6
      part_block_insert = part_block.copy()
      part_block_offset_x = self.block_offset_x - part_block.dxf.insert.x
      part_block_offset_y = self.block_offset_y - part_block.dxf.insert.y

      part_block_insert.translate(part_block_offset_x, part_block_offset_y, 0)

      values = {
        "MARK": part.tag,
        "RM": "0",
        "CODE": "TAG-CTRL+F-" + part.tag,
        "TYPE": "STEELWORK",
        "REF-DWG": "REFERENCE DRAWING",
        "WEIGHT": part.unit_weight,
        "DESCL1": part.description,
        "QTY": part.quantity,
        "TOTAL": part.total_weight,
      }

      for tag, valor in values.items():

        attrib = part_block_insert.get_attrib(tag)

        if attrib:
            attrib.dxf.text = str(valor)

      self.msp_dxf.add_entity(part_block_insert)

  # Função para adicionar Text ao DXF
  def add_text(self, text, position_insert, attributes_text):
    attributes_text["insert"] = position_insert
    self.msp_dxf.add_text(text=text, dxfattribs=attributes_text)

  # Adiciona a Tag e o quadrado da Tag ao DXF
  def addTagDraw(self, text, textPosition, retanglePoints):
    tagPropertiesNotes = {
      "height": 3,
      "style": "ROMAND",
      "rotation": 0,
      "width": 0.8,
      "layer" : "CONTORNI",
    }

    self.msp_dxf.add_text(text=text, dxfattribs=tagPropertiesNotes).set_placement(textPosition, align=TextEntityAlignment.CENTER)
        
    self.msp_dxf.add_lwpolyline(
      points = retanglePoints,
      dxfattribs={
        "layer" : "SOTTILI"
      }
    )

  def add_general_notes(self, drawing_code, part_list):
    general_properties_notes = { 
      "height": 3,  
      "style": "Standard",
      "rotation": 0,
      "layer" : "TESTI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    general_properties_notes_contorni = { 
      "height": 3,  
      "style": "Standard",
      "rotation": 0,
      "layer" : "CONTORNI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    self.add_text("- ALL PIECES TO BE MARKED WITH:", (self.block_offset_x + 35.333333, self.block_offset_y, 0), general_properties_notes)
    rectanglePointsTag = [
      (self.block_offset_x + 106, self.block_offset_y + 5, 0), 
      (self.block_offset_x + 180, self.block_offset_y + 5, 0), 
      (self.block_offset_x + 180, self.block_offset_y - 2, 0), 
      (self.block_offset_x + 106, self.block_offset_y - 2, 0), 
      (self.block_offset_x + 106, self.block_offset_y + 5, 0)
    ]
    self.addTagDraw("TAG-CTRL+F-" + drawing_code, (self.block_offset_x + 143, self.block_offset_y, 0), rectanglePointsTag)
    self.block_offset_y += 5

    notes_weight_part_block = self.doc_dxf.blocks.get('NOTE-PARTS-WEIGHT')
    notes_weight_part_block_insert = self.msp_dxf.add_blockref(notes_weight_part_block.name, insert=(self.block_offset_x + 180, self.block_offset_y, 0))

    values = {
      "PIECES-REQUIRED1": str(part_list[0].quantity),
      "PIECES-REQUIRED2": f"{part_list[0].quantity:02d}",
      "TOTAL-WEIGHT": part_list[0].total_weight + " kg",
      "UNIT-WEIGHT": part_list[0].unit_weight + " kg",
    }

    notes_weight_part_block_insert.add_auto_attribs(values)
    self.block_offset_y += 18

class DrawingList:
  def __init__(self, sheet, sheet_parts, sheet_material):
    self.drawing_code_list = sheet
    self.list_part = PartList(sheet_parts)
    self.list_material = MaterialList(sheet_material)

class PartList:
  def __init__(self, sheet_parts):
    self.part = self.read_sheet_part(sheet_parts)
    self.approx_total_weight = self.get_approx_total_weight()

  # Lê informações da Planilha de Peças
  def read_sheet_part(self, sheet_parts):
    part_list = []

    for row in range(2, sheet_parts.max_row + 1):
      cell_value = sheet_parts.cell(row=row, column=1).value

      if cell_value == None:
        break

      tag = sheet_parts.cell(row=row, column=1).value
      description = sheet_parts.cell(row=row, column=6).value or sheet_parts.cell(row=row, column=7).value or sheet_parts.cell(row=row, column=5).value
      quantity = sheet_parts.cell(row=row, column=4).value
      material = sheet_parts.cell(row=row, column=8).value
      unit_weight = f"{sheet_parts.cell(row=row, column=11).value:.1f}"
      total_weight = f"{sheet_parts.cell(row=row, column=12).value:.1f}"

      part_list.append(Part(tag, description, quantity, material, unit_weight, total_weight))

    return part_list
  
  # Lê informações do Peso das Peças e Soma dando o Peso total arrendonda sem casa decimal
  def get_approx_total_weight(self):
    approx_total_weight = 0.0
    
    for part in self.part:
      approx_total_weight = approx_total_weight + float(part.total_weight)

    approx_total_weight = str(round(approx_total_weight))
    
    return approx_total_weight
    
class MaterialList:
  def __init__(self, sheet_material):
    self.materials = self.read_sheet_material(sheet_material)
    self.material_meter = None
    self.material_square_meter = None
    self.get_material_for_meter()

  # Lê informações da Planilha de Materiais
  def read_sheet_material(self, sheet_material):
    material_list = []

    for row in range(2, sheet_material.max_row + 1):
      cell_value = sheet_material.cell(row=row, column=12).value
      if cell_value == None: break

      normalized_quantity  = sheet_material.cell(row=row, column=15).value
      normalized_quantity  = f"{normalized_quantity :.1f}" if isinstance(normalized_quantity , (int, float)) else normalized_quantity 
      
      position = sheet_material.cell(row=row, column=12).value
      description = sheet_material.cell(row=row, column=13).value
      unit = sheet_material.cell(row=row, column=14).value
      quantity = normalized_quantity 
      composition = sheet_material.cell(row=row, column=16).value
      total_weight = f"{sheet_material.cell(row=row, column=17).value:.1f}"

      material_list.append(Material(position, description, unit, quantity, composition, total_weight))

    return material_list
  
  # Lê informações dos Materiais pelo Metro e Metro Quadrado para indicar o Material das Chapas e Material dos Perfil
  def get_material_for_meter(self):
    material_meter = set()
    material_square_meter = set()

    # Itere pela lista de dados e concatene o material com base na unidade
    for material in self.materials:
      if material.unit == 'm':
        material_meter.add(material.composition)
      elif material.unit == 'm²':
        material_meter.add(material.composition)

    material_meter_join = " / ".join(material_meter)
    material_square_meter_join = " / ".join(material_square_meter)

    self.material_meter = material_meter_join
    self.material_square_meter = material_square_meter_join

class Part:
  def __init__(self, tag, description, quantity, material, unit_weight, total_weight):
    self.tag = tag
    self.description = description
    self.quantity = quantity
    self.material = material
    self.unit_weight = unit_weight
    self.total_weight = total_weight

class Material:
  def __init__(self, position, description, unit, quantity, composition, total_weight):
    self.position = position
    self.description = description
    self.unit = unit
    self.quantity = quantity
    self.composition = composition
    self.total_weight = total_weight