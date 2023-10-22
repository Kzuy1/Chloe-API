import ezdxf
from ezdxf.enums import TextEntityAlignment
import openpyxl
import os
import zipfile

def lista(fullPath) :
  workbook = openpyxl.load_workbook(fullPath, data_only= True)
  fatherPath = os.path.abspath(".")
  fileName = os.path.splitext(os.path.basename(fullPath))[0]

  zipFile = fatherPath + '/list/'+ fileName +'.zip'
  with zipfile.ZipFile(zipFile, 'w', compression=zipfile.ZIP_DEFLATED) as myzip:
   myzip.write(fatherPath + '/BLOCO-BRANCO.dxf', arcname='BLOCO-BRANCO.dxf')

  def writeFiles(elemento, index):
    worksheetMaterial = workbook[elemento + ' - MATERIAL']
    MaterialList = []

    for row in range(2, worksheetMaterial.max_row + 1):
      cell_value = worksheetMaterial.cell(row=row, column=12).value
      if cell_value == None:
        break

      QTDE = worksheetMaterial.cell(row=row, column=15).value
      QTDE = f"{QTDE:.1f}" if isinstance(QTDE, (int, float)) else QTDE
      
      material = {
        'POS' : worksheetMaterial.cell(row=row, column=12).value,
        'descricao' : worksheetMaterial.cell(row=row, column=13).value,
        'unidade' : worksheetMaterial.cell(row=row, column=14).value,
        'QTDE' : QTDE,
        'material' : worksheetMaterial.cell(row=row, column=16).value,
        'PesoTotal' : f"{worksheetMaterial.cell(row=row, column=17).value:.1f}"
      }
      MaterialList.append(material)

    worksheetPeca = workbook[elemento + ' - PEÇAS']
    PecaList = []

    for row in range(2, worksheetPeca.max_row + 1):
      cell_value = worksheetPeca.cell(row=row, column=1).value
      if cell_value == None:
        break
      peca = {
        'cod' : '...' + worksheetPeca.cell(row=row, column=2).value,
        'descricao' : worksheetPeca.cell(row=row, column=6).value or worksheetPeca.cell(row=row, column=7).value or worksheetPeca.cell(row=row, column=5).value,
        'QTDE' : worksheetPeca.cell(row=row, column=4).value,
        'material' : worksheetPeca.cell(row=row, column=8).value,
        'PesoUnit' : f"{worksheetPeca.cell(row=row, column=11).value:.1f}",
        'PesoTotal' : f"{worksheetPeca.cell(row=row, column=12).value:.1f}"
      }
      PecaList.append(peca)

    doc = ezdxf.readfile(fatherPath + "/list/" + fileName + '.dxf')
    msp = doc.modelspace()

    posForYBlock = 0
    posForXBlock = 270 * index

    msp.add_lwpolyline([(posForXBlock, 0, 0), (posForXBlock, 574, 0), (posForXBlock+180, 574, 0), (posForXBlock+180, 0, 0), (posForXBlock, 0, 0)])

    blocoLegenda = doc.blocks.get('REDECAM-TITOLO-TAVOLA')
    if blocoLegenda :
      msp.add_blockref(blocoLegenda.name, insert=(posForXBlock + 180, posForYBlock, 0))
      posForYBlock += 148
    
    blocoRevisao = doc.blocks.get('REDE-DISTINTA-REVISIONE')
    if blocoRevisao :
      for i in range(4) :
        blocoRevisaoInsert = msp.add_blockref(blocoRevisao.name, insert=(posForXBlock + 180, posForYBlock, 0))
        blocoRevisaoInsert.add_attrib("REV-N", i)
        posForYBlock += 7

    bloco_material = doc.blocks.get('EMB_LISTA_DE_MATERIAL')
    if bloco_material :
      posForYBlock += 13
      for objeto in MaterialList:
          posForYBlock += 7
          insert = msp.add_blockref(bloco_material.name, insert=(posForXBlock, posForYBlock, 0))
          insert.add_attrib("POSICAO", objeto["POS"])
          insert.add_attrib("DESCRICAO", objeto["descricao"])
          insert.add_attrib("UNIDADE", objeto["unidade"])
          insert.add_attrib("QUANTIDADE", objeto["QTDE"])
          insert.add_attrib("MATERIAL", objeto["material"])
          insert.add_attrib("PESO", objeto["PesoTotal"])

    msp.add_blockref('EMB_LISTA_DE_MATERIAL_DESCRITIVO_ING', insert=(posForXBlock, 182, 0))
    msp.add_blockref('EMB_LISTA_DE_MATERIAL_INFO_ING', insert=(posForXBlock, posForYBlock, 0))

    blocoBranco = doc.blocks.get('REDECAM-DISTINTA_monolingua')
    if blocoBranco :
      posForYBlock += 7
      for objeto in PecaList:
          posForYBlock += 6
          insert = msp.add_blockref(blocoBranco.name, insert=(posForXBlock, posForYBlock, 0))
          insert.add_attrib("MARCA", objeto["cod"])
          insert.add_attrib("DESCRIZIONE-IT", objeto["descricao"])
          insert.add_attrib("DESCRIZIONE-IN-R1", objeto["descricao"])
          insert.add_attrib("QUANTITA'", objeto["QTDE"])
          insert.add_attrib("MATERIALE", objeto["material"])
          insert.add_attrib("PESO-CAD", objeto["PesoUnit"])
          insert.add_attrib("TOTALE", objeto["PesoTotal"])
          insert.add_attrib("LARGHEZZA", 0)
          insert.add_attrib("PROFONDITA'", 0)
          insert.add_attrib("ALTEZZA", 0)
          insert.add_attrib("CICLO-VERN-INT", 0)
          insert.add_attrib("CICLO-VERN-EST", 0)
          insert.add_attrib("VERNICIATURA-INT", 0)
          insert.add_attrib("VERNICIATURA-EST", 0)

    #Pega o Peso Total do Desenho
    total_weight = 0.0
    for item in PecaList:
        total_weight += float(item['PesoTotal'])
    total_weight = str(round(total_weight))

    #Pega Informações do Material
    materiais_metro = set()
    materiais_metro_quadrado = set()

    # Itere pela lista de dados e concatene o material com base na unidade
    for item in MaterialList:
        if item['unidade'] == 'm':
            materiais_metro.add(item['material'])
        elif item['unidade'] == 'm²':
            materiais_metro_quadrado.add(item['material'])

    material_metro = " / ".join(materiais_metro)
    material_metro_quadrado = " / ".join(materiais_metro_quadrado)

    #Insere as Notas
    attrib_properties = { 
      "height": 3,  
      "style": "Standard",
      "rotation": 0,
      "layer" : "TESTI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    def adicionar_texto(texto, insert):
      attrib_properties["insert"] = insert
      msp.add_text(text=texto, dxfattribs=attrib_properties)
    
    posForYBlock += 12
    adicionar_texto("  OTHERWISE WHERE INDICATED.", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 5
    adicionar_texto("- ALL BEND RADIUS ARE EQUAL TO THE VALUE OF THE THICKNESS OF THE SHEET", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("  EXCEPT THOSE SHOWED BY #, ASSEMBLED OR WELDED TO THE PART", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 5
    adicionar_texto("- THE BOLTS AND NUTS IN THE TABLE MUST BE DISPATCHED LOOSE", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("- ALL IDENTICAL PARTS MUST HAVE THE SAME MARK", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("- ALL COMPONENTS TO BE FORWARDED LOOSE MUST BE IDENTIFIED BY A MARK TAG", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("  ON THE PART INDICATED IN THE DRAWING", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 5
    adicionar_texto("- THE COMPONENTS SHOWED BY # MUST BE ASSEMBLED IN THE WORKSHOP", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("- %%UIMPORTANT:%%U  PRE-ASSEMBLY IN WORK-SHOP", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("-     = BOLT INDICATED IN OTHER DRAWING", (posForXBlock + 5, posForYBlock, 0))
    msp.add_circle(
      center= (posForXBlock + 14, posForYBlock+1.5, 0),
      radius= 3,
      dxfattribs={
        "layer" : "NASCOSTE"
      }
    )
    msp.add_line(
      start=(posForXBlock + 11.879, posForYBlock+3.621, 0),
      end=(posForXBlock + 16.121, posForYBlock-0.621, 0),
      dxfattribs={
        "layer" : "NASCOSTE" 
      }
    )
    msp.add_line(
      start=(posForXBlock + 11.879, posForYBlock-0.621, 0),
      end=(posForXBlock + 16.121, posForYBlock+3.621, 0),
      dxfattribs={
        "layer" : "NASCOSTE" 
      }
    )

    posForYBlock += 7
    adicionar_texto("- FOR CONSTRUCTION AND SUPPLY GENERAL NOTES, SEE SPECIFICATION \"SR-R1-01\"", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("- REFERENCE DRAWINGS: ___ ÷ ___", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    adicionar_texto("- FOR ASSEMBLY DRAWING SEE DWG No.:  ___", (posForXBlock + 5, posForYBlock, 0))

    posForYBlock += 7
    msp.add_text(text="- TOTAL WEIGHT:  "+total_weight+" kg approx", 
                 dxfattribs={
                    "height": 3,  
                    "style": "ROMAND",
                    "rotation": 0,
                    "layer" : "CONTORNI",
                    "width": 0.8,
                    "insert": (posForXBlock + 5, posForYBlock, 0)
                 })
    
    posForYBlock += 7
    if(material_metro != ""):
      adicionar_texto("- PROFILES MATERIAL:  " + material_metro, (posForXBlock + 5, posForYBlock, 0))
      posForYBlock += 7
    if(material_metro_quadrado != ""):
      adicionar_texto("- SHEET MATERIAL:  " + material_metro_quadrado, (posForXBlock + 5, posForYBlock, 0))
      posForYBlock += 7
    
    adicionar_texto("- ALL PIECES TO BE MARKED WITH:", (posForXBlock + 5, posForYBlock, 0))

    msp.add_text(text="AA-BX-XX/...",
                 dxfattribs={
                    "height": 3,
                    "style": "ROMAND",
                    "rotation": 0,
                    "width": 0.8,
                    "layer" : "CONTORNI",
                 }).set_placement((posForXBlock + 95, posForYBlock, 0), align=TextEntityAlignment.CENTER)
    msp.add_lwpolyline(
      points=[(posForXBlock + 78, posForYBlock+5, 0), (posForXBlock + 112, posForYBlock+5, 0), (posForXBlock + 112, posForYBlock-2, 0), (posForXBlock + 78, posForYBlock-2, 0), (posForXBlock + 78, posForYBlock+5, 0)],
      dxfattribs={
        "layer" : "SOTTILI"
      }
    )
    
    posForYBlock += 9
    msp.add_text(text="%%uANNOTATIONS:%%u", 
              dxfattribs={
                "height": 4,  
                "style": "ROMAND",
                "rotation": 0,
                "layer" : "NOTE",
                "width": 0.8,
                "insert": (posForXBlock + 5, posForYBlock, 0)
              })

    # Exiba o material em diferentes unidades

    mtext = msp.add_mtext(
      text= "\\A1;" + elemento,
      dxfattribs={
        "insert": (posForXBlock + 90, 599, 0),  
        "char_height": 25,  
        "width": 0,  
        "style": "Standard",
        "rotation": 0,
        "color": 4,
        "layer" : "QUOTE",
        "attachment_point": 5
      }
      )
    
    doc.save()


  planilhas = workbook.worksheets
  nomes_planilhas = set()

  for planilha in planilhas[1:]:
      primeiro_valor = planilha.title.split('-')[0].strip()
      nomes_planilhas.add(primeiro_valor)

  nomes_planilhas = list(nomes_planilhas)
  nomes_planilhas.sort()
  
  dxfPath = fatherPath + "/list/" + fileName + '.dxf'
  CreateFile = ezdxf.readfile(fatherPath + '/BLOCO-BRANCO.dxf')
  CreateFile.saveas(dxfPath)

  for index, elemento in enumerate(nomes_planilhas): 
    writeFiles(elemento, index)
  
  with zipfile.ZipFile(zipFile, 'a', compression=zipfile.ZIP_DEFLATED) as myzip:
    myzip.write(dxfPath, arcname=fileName + '.dxf')
  
  #os.remove(dxfPath)

  return zipFile