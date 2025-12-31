import os
import zipfile
import ezdxf
from openpyxl import load_workbook
from ezdxf.enums import TextEntityAlignment

def clear_folder(folder_path):
    if os.path.exists(folder_path):
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if os.path.isfile(file_path):
                os.remove(file_path)

def import_attributes_from_xlsx(xlsx_file, zip_file):
    temp_dir = os.path.join(os.path.dirname(__file__), "temp")
    
    uploads_folder = os.path.join(temp_dir, "uploads")
    extracted_folder = os.path.join(temp_dir, "extracted")
    edited_folder = os.path.join(temp_dir, "edited")
    output_folder = os.path.join(temp_dir, "output")

    os.makedirs(uploads_folder)
    os.makedirs(extracted_folder)
    os.makedirs(edited_folder)
    os.makedirs(output_folder)

    xlsx_path = os.path.join(uploads_folder, xlsx_file.filename)
    zip_path = os.path.join(uploads_folder, zip_file.filename)
    xlsx_file.save(xlsx_path)
    zip_file.save(zip_path)

    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extracted_folder)

    workbook = load_workbook(xlsx_path)
    worksheet = workbook.active
    header = [cell.value for cell in worksheet[1]]
    idx_name = header.index("PartName")
    idx_thickness = header.index("Thickness")
    idx_material = header.index("Material")
    idx_quantity = header.index("Quantity")

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        part_name = str(row[idx_name])
        thickness = str(row[idx_thickness]).replace('.', ',')
        material = row[idx_material]
        quantity = row[idx_quantity]

        file_name = f'{part_name}.dxf'
        input_path = os.path.join(extracted_folder, file_name)
        output_path = os.path.join(edited_folder, file_name)

        if not os.path.exists(input_path):
            print(f'[Aviso] Arquivo não encontrado: {file_name}')
            continue

        try:
            doc = ezdxf.readfile(input_path)
            msp = doc.modelspace()

            if "Administração" not in doc.layers:
                doc.layers.new(name='Administração', dxfattribs={'color': 1})

            texts = [
                f'Part Name {part_name}',
                f'Thickness {thickness}',
                f'Material {material}',
                f'Quantity {quantity}',
            ]

            y_offset = -15
            for text in texts:
                msp.add_text(
                    text,
                    dxfattribs={
                        'layer': 'Administração',
                        'height': 10,
                    }
                ).set_placement((0, y_offset), align=TextEntityAlignment.LEFT)
                y_offset -= 12

            doc.saveas(output_path)
        except Exception as e:
            print(f'[Erro] {file_name}: {e}')

    clear_folder(uploads_folder)
    clear_folder(extracted_folder)

    output_path = os.path.join(output_folder, f"resultado.zip")

    zip_final_path = os.path.join(output_folder, "resultado.zip")
    with zipfile.ZipFile(zip_final_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_name in os.listdir(edited_folder):
            file_path = os.path.join(edited_folder, file_name)
            zipf.write(file_path, arcname=file_name)

    clear_folder(edited_folder)

    return zip_final_path