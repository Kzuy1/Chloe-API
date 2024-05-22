import ezdxf
from ezdxf.enums import TextEntityAlignment
from pymongo import MongoClient
from json import load
import openpyxl
import os

class listToDXF:
  def __init__(self, fullPath):
    self.fullPath = fullPath
    self.fatherPath = os.path.abspath(".")
    self.fileName = os.path.splitext(os.path.basename(self.fullPath))[0]
    self.workbook = openpyxl.load_workbook(fullPath, data_only=True)
    self.drawnList = self.readSheetsNames()
    self.projectCodeDraw = self.getProjectCode()
    self.drawnLanguage = self.findStandardLanguage()
    self.targetDXFFile = self.createDXFFile()
    self.docDXF = ezdxf.readfile(self.targetDXFFile)
    self.mspDXF = self.docDXF.modelspace()
    self.posForXBlock = 0
    self.posForYBlock = 0
    
    # Loop para adicionar todas as Listas no 
    for index, drawnList in enumerate(self.drawnList):
      # Determina a posição dos Textos, Blocos etc.
      self.posForXBlock = 270 * index
      self.posForYBlock = 0
    
      # Adiciona o Bloco de Lengenda ao DXF
      self.addDrawingSubtitle()

      # Adiciona o Blocos de Revisão ao DXF
      self.addRevisionSubtitle()

      # Adiciona a Lista de Material ao DXF
      self.addMaterialList(drawnList.listMateiral.materials)

      # Adiciona o Bloco Descritivo de Material ao DXF
      self.addDescriptionMaterialList()

      # Adiciona os Blocos de Peças ao DXF
      self.addPieceList(drawnList.listPiece.pieces)

      # Adiciona as Notas Gerais
      drawApproxTotalWeight = drawnList.listPiece.approxTotalWeight
      drawMaterialMeter = drawnList.listMateiral.materialMeter
      drawMaterialSquareMeter = drawnList.listMateiral.materialSquareMeter
      if self.drawnLanguage == 'brazil':
        self.addGeneralNotesPTBR(drawApproxTotalWeight, drawMaterialMeter, drawMaterialSquareMeter)
      else :
        self.addGeneralNotesENUS(drawApproxTotalWeight, drawMaterialMeter, drawMaterialSquareMeter)

      # Adiciona um retangulo em volta das informações, para dividir em vários informações
      self.mspDXF.add_lwpolyline([(self.posForXBlock, 0, 0), (self.posForXBlock, 574, 0), (self.posForXBlock+180, 574, 0), (self.posForXBlock+180, 0, 0), (self.posForXBlock, 0, 0)])
      
      # Adiciona um indentificador do Codigo do Desenho em cima do Retangulo
      self.mspDXF.add_mtext(
        text= "\\A1;" + drawnList.drawnCodeList,
        dxfattribs={
          "insert": (self.posForXBlock + 90, 599, 0),  
          "char_height": 25,  
          "width": 0,  
          "style": "Standard",
          "rotation": 0,
          "color": 4,
          "layer" : "QUOTE",
          "attachment_point": 5
        }
      )
      # Salva o DXF
      self.docDXF.save()
      
  # Função para pegar as Listas do Desenho
  def readSheetsNames(self):
    sheetsNamesCollection = set()

    for sheet in self.workbook.worksheets[1:]:
      name = sheet.title.split('-')[0].strip()
      sheetsNamesCollection.add(name)

    sheetsNamesArray = list(sheetsNamesCollection)
    sheetsNamesArray.sort()

    drawingListArray = []
    for sheet in sheetsNamesArray:
      drawingListArray.append(drawingList(sheet, self.workbook[sheet + ' - PEÇAS'], self.workbook[sheet + ' - MATERIAL']))
    
    return drawingListArray
  
  # Função para determinar código do Projeto
  def getProjectCode(self):
    targetSheet = self.workbook.worksheets[0]
    collumnPieceNumber = targetSheet['E']
    projectCode = None
    highestCount = 0

    countOccurrences = {}
    for cell in collumnPieceNumber[1:]:
      if isinstance(cell.value, str):
        cell_value = cell.value[:7]
        countOccurrences[cell_value] = countOccurrences.get(cell_value, 0) + 1

    for value, count in countOccurrences.items():
      if count > highestCount:
        projectCode = value
        highestCount = count

    return projectCode

  # Função para pegar a Standard do Banco de Dados.
  def findStandardLanguage(self):
    # Conectar ao MongoDB
    with open('config.json', 'r') as fileConfig:
      config = load(fileConfig)

    # Procura no Banco de Dados - EMBRATECENG na Collection - Projects a informações de Standard do código do Projeto
    mongoClient = MongoClient(config['mongoUrl'])
    mongoDatabase = mongoClient.get_database('EMBRATECENG')
    collectionProjects = mongoDatabase['projects']
    result = collectionProjects.find_one({"cod": self.projectCodeDraw})
    mongoClient.close()

    # Caso não ache use o Brazil como referencia
    standarProject = result['standard'] if result else 'brazil'
  
    return standarProject
    
  # Função para copiar o Template e criar um novo arquivo para Edição
  def createDXFFile(self):
    targetDXFFile = self.fatherPath + "/list/" + self.fileName + '.dxf'

    # Caso arquivo exista ele irá remover
    if os.path.isfile(targetDXFFile):
      os.remove(targetDXFFile)

    CreateFile = ezdxf.readfile(self.fatherPath + '/BLOCO-BRANCO.dxf')
    CreateFile.saveas(targetDXFFile)
    
    return targetDXFFile
  
  # Função para adicionar o Bloco de Lengenda ao Desenho
  def addDrawingSubtitle(self):
    subtitleBlock = self.docDXF.blocks.get('REDECAM-TITOLO-TAVOLA')
    if not subtitleBlock : return

    self.mspDXF.add_blockref(subtitleBlock.name, insert=(self.posForXBlock + 180, self.posForYBlock, 0))
    self.posForYBlock += 148

  # Função para adicionar o Bloco de Revisão ao Desenho
  def addRevisionSubtitle(self):
    revisionBlock = self.docDXF.blocks.get('REDECAM_REVISION')
    if not revisionBlock : return

    for i in range(4) :
      revisionBlockInsert = self.mspDXF.add_blockref(revisionBlock.name, insert=(self.posForXBlock + 180, self.posForYBlock, 0))
      revisionBlockInsert.add_attrib("REV-N", i)
      self.posForYBlock += 7

  # Função para adicionar a Lista de Material ao Blocos de Material e adicionar ao DXF
  def addMaterialList(self, listMaterial):
    materialBlock = self.docDXF.blocks.get('EMB_LISTA_DE_MATERIAL_v0.2')
    if not materialBlock : return

    self.posForYBlock += 13
    for material in listMaterial:
      self.posForYBlock += 7
      materialBlockInsert = self.mspDXF.add_blockref(materialBlock.name, insert=(self.posForXBlock, self.posForYBlock, 0))

      values = {
        "POSICAO": material.position,
        "DESCRICAO": material.description,
        "UNIDADE": material.unit,
        "QUANTIDADE": material.quantity,
        "MATERIAL": material.composition,
        "PESO": material.totalWeight
      }

      materialBlockInsert.add_auto_attribs(values)
  
  # Função para adicionar o Bloco Descritivo de Material ao DXF
  def addDescriptionMaterialList(self):
    if self.drawnLanguage == 'brazil' :
      descriptionMaterialBlock = self.docDXF.blocks.get('EMB_LISTA_DE_MATERIAL_DESCRITIVO_v0.2')
      infoMaterialBlock = self.docDXF.blocks.get('EMB_LISTA_DE_MATERIAL_INFO_v0.2')
    else :
      descriptionMaterialBlock = self.docDXF.blocks.get('EMB_LISTA_DE_MATERIAL_DESCRITIVO_ING_v0.2')
      infoMaterialBlock = self.docDXF.blocks.get('EMB_LISTA_DE_MATERIAL_INFO_ING_v0.2')

    self.mspDXF.add_blockref(descriptionMaterialBlock.name, insert=(self.posForXBlock, 182, 0))
    self.mspDXF.add_blockref(infoMaterialBlock.name, insert=(self.posForXBlock, self.posForYBlock, 0))

  # Adiciona os Blocos de Peças ao DXF
  def addPieceList(self, listPiece):
    pieceBlock = self.docDXF.blocks.get('REDECAM-DISTINTA_monolingua')
    if not pieceBlock : return

    self.posForYBlock += 7
    for piece in listPiece:
      self.posForYBlock += 6
      pieceBlockInsert = self.mspDXF.add_blockref(pieceBlock.name, insert=(self.posForXBlock, self.posForYBlock, 0))

      values = {
        "MARCA": piece.cod,
        "DESCRIZIONE-IT": piece.description,
        "DESCRIZIONE-IN-R1": piece.description,
        "QUANTITA'": piece.quantity,
        "MATERIALE": piece.material,
        "PESO-CAD": piece.unitWeight,
        "TOTALE": piece.totalWeight,
        "LARGHEZZA": 0,
        "PROFONDITA'": 0,
        "ALTEZZA": 0,
        "CICLO-VERN-INT": 0,
        "CICLO-VERN-EST": 0,
        "VERNICIATURA-INT": 0,
        "VERNICIATURA-EST": 0
      }

      pieceBlockInsert.add_auto_attribs(values)

  # Função para adicionar mText ao DXF
  def addMtext(self, text, positionInsert, attributesMtext):
    attributesMtext["insert"] = positionInsert
    self.mspDXF.add_text(text=text, dxfattribs=attributesMtext)

  # Adiciona a Tag e o quadrado da Tag ao DXF
  def addTagDraw(self, textPosition, retanglePoints):
    tagPropertiesNotes = {
      "height": 3,
      "style": "ROMAND",
      "rotation": 0,
      "width": 0.8,
      "layer" : "CONTORNI",
    }

    self.mspDXF.add_text(text="AA-BX-XX/...", dxfattribs=tagPropertiesNotes).set_placement(textPosition, align=TextEntityAlignment.CENTER)
        
    self.mspDXF.add_lwpolyline(
      points = retanglePoints,
      dxfattribs={
        "layer" : "SOTTILI"
      }
    )

  # Adiciona Notas Gerais ao DXF
  def addGeneralNotesENUS(self, approxTotalWeight, materialMeter, materialSquareMeter):
    generalPropertiesNotes = { 
      "height": 3,  
      "style": "Standard",
      "rotation": 0,
      "layer" : "TESTI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    approxTotalWeighttribsPropertiesNotes = {
      "height": 3,  
      "style": "ROMAND",
      "rotation": 0,
      "layer" : "CONTORNI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    annotationsPropertiesNotes = {
      "height": 4,  
      "style": "ROMAND",
      "rotation": 0,
      "layer" : "NOTE",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    self.posForYBlock += 12
    self.addMtext("  EXCEPT THOSE SHOWED BY #, ASSEMBLED OR WELDED TO THE PART", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 5
    self.addMtext("- THE BOLTS AND NUTS IN THE TABLE MUST BE DISPATCHED LOOSE", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- ALL IDENTICAL PARTS MUST HAVE THE SAME MARK", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- ALL COMPONENTS TO BE FORWARDED LOOSE MUST BE IDENTIFIED BY A MARK TAG", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("  ON THE PART INDICATED IN THE DRAWING", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 5
    self.addMtext("- THE COMPONENTS SHOWED BY # MUST BE ASSEMBLED IN THE WORKSHOP", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
    
    self.posForYBlock += 7
    self.addMtext("- FOR CONSTRUCTION AND SUPPLY GENERAL NOTES, SEE SPECIFICATION \"SR-R1-01\"", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- %%UIMPORTANT:%%U  PRE-ASSEMBLY IN WORK-SHOP", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    boltBlock = self.docDXF.blocks.get('1418104')
    self.mspDXF.add_blockref(boltBlock.name, insert=(self.posForXBlock + 5, self.posForYBlock, 0))

    self.posForYBlock += 7
    self.addMtext("- REFERENCE DRAWINGS No.: ___ ÷ ___", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- FOR ASSEMBLY DRAWING SEE DWG No.: ___", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- TOTAL WEIGHT, APPROX: " + approxTotalWeight + " kg", (self.posForXBlock + 5, self.posForYBlock, 0), approxTotalWeighttribsPropertiesNotes)

    self.posForYBlock += 7
    if(materialMeter != ""):
      self.addMtext("- PROFILES MATERIAL: " + materialMeter, (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
      self.posForYBlock += 7

    if(materialSquareMeter != ""):
      self.addMtext("- SHEET MATERIAL: " + materialSquareMeter, (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
      self.posForYBlock += 7
    
    self.addMtext("- ALL PIECES TO BE MARKED WITH:", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
    rectanglePointsTag = [
      (self.posForXBlock + 78, self.posForYBlock + 5, 0), 
      (self.posForXBlock + 112, self.posForYBlock + 5, 0), 
      (self.posForXBlock + 112, self.posForYBlock - 2, 0), 
      (self.posForXBlock + 78, self.posForYBlock - 2, 0), 
      (self.posForXBlock + 78, self.posForYBlock + 5, 0)
    ]
    self.addTagDraw((self.posForXBlock + 95, self.posForYBlock, 0), rectanglePointsTag)

    self.posForYBlock += 8
    self.addMtext("%%uANNOTATIONS:%%u", (self.posForXBlock + 5, self.posForYBlock, 0), annotationsPropertiesNotes)

  def addGeneralNotesPTBR(self, approxTotalWeight, materialMeter, materialSquareMeter):
    generalPropertiesNotes = { 
      "height": 3,  
      "style": "Standard",
      "rotation": 0,
      "layer" : "TESTI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    approxTotalWeighttribsPropertiesNotes = {
      "height": 3,  
      "style": "ROMAND",
      "rotation": 0,
      "layer" : "CONTORNI",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    annotationsPropertiesNotes = {
      "height": 4,  
      "style": "ROMAND",
      "rotation": 0,
      "layer" : "NOTE",
      "width": 0.8,
      "insert": (0, 0, 0)
    }

    self.posForYBlock += 12
    self.addMtext("  EXCETO AQUELES MOSTRADOS POR #, MONTADOS OU SOLDADOS À PEÇA", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 5
    self.addMtext("- OS PARAFUSOS E PORCAS NA TABELA DEVEM SER ENVIADOS SOLTOS", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- TODAS AS PEÇAS IGUAIS DEVEM TER A MESMA MARCAÇÃO", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("  COM UMA MARCA/ETIQUETA", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 5
    self.addMtext("- TODOS OS COMPONENTES A SEREM ENCAMINHADOS SOLTOS DEVEM SER IDENTIFICADOS", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("  NA PARTE INDICADA NO DESENHO", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 5
    self.addMtext("- OS COMPONENTES MOSTRADOS COM # DEVEM SER MONTADOS NA FABRICAÇÃO", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- PARA NOTAS GERAIS DE CONSTRUÇÃO E FORNECIMENTO, VER ESPECIFICAÇÃO \"SR-R1-01\"", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- %%UIMPORTANTE:%%U  PRÉ-MONTAGEM NA FABRICAÇÃO", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    boltBlock = self.docDXF.blocks.get('A$C244a8b43')
    self.mspDXF.add_blockref(boltBlock.name, insert=(self.posForXBlock + 5, self.posForYBlock, 0))

    self.posForYBlock += 7
    self.addMtext("- DESENHOS DE REFERÊNCIA N°: ___ ÷ ___", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- PARA DESENHO DE MONTAGEM VER DESENHO N°: ___", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)

    self.posForYBlock += 7
    self.addMtext("- PESO TOTAL, APROX: " + approxTotalWeight + " kg", (self.posForXBlock + 5, self.posForYBlock, 0), approxTotalWeighttribsPropertiesNotes)

    self.posForYBlock += 7
    if(materialMeter != ""):
      self.addMtext("- MATERIAL DO PERFIL: " + materialMeter, (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
      self.posForYBlock += 7

    if(materialSquareMeter != ""):
      self.addMtext("- MATERIAL DAS CHAPAS: " + materialSquareMeter, (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
      self.posForYBlock += 7
    
    self.addMtext("- TODAS AS PEÇAS DEVEM SER MARCADAS COM:", (self.posForXBlock + 5, self.posForYBlock, 0), generalPropertiesNotes)
    rectanglePointsTag = [
      (self.posForXBlock + 103, self.posForYBlock + 5, 0), 
      (self.posForXBlock + 137, self.posForYBlock + 5, 0), 
      (self.posForXBlock + 137, self.posForYBlock - 2, 0), 
      (self.posForXBlock + 103, self.posForYBlock - 2, 0), 
      (self.posForXBlock + 103, self.posForYBlock + 5, 0)
    ]
    self.addTagDraw((self.posForXBlock + 120, self.posForYBlock, 0), rectanglePointsTag)

    self.posForYBlock += 8
    self.addMtext("%%uNOTAS:%%u", (self.posForXBlock + 5, self.posForYBlock, 0), annotationsPropertiesNotes)

class drawingList:
  def __init__(self, sheet, sheetPiece, sheetMaterial):
    self.drawnCodeList = sheet
    self.listPiece = pieceList(sheetPiece)
    self.listMateiral = materialList(sheetMaterial)

class pieceList:
  def __init__(self, sheetPiece):
    self.pieces = self.readSheetPiece(sheetPiece)
    self.approxTotalWeight = self.getApproxTotalWeight()

  # Lê informações da Planilha de Peças
  def readSheetPiece(self, sheetPiece):
    pieceList = []

    for row in range(2, sheetPiece.max_row + 1):
      cell_value = sheetPiece.cell(row=row, column=1).value

      if cell_value == None:
        break

      cod = '...' + sheetPiece.cell(row=row, column=2).value
      description = sheetPiece.cell(row=row, column=6).value or sheetPiece.cell(row=row, column=7).value or sheetPiece.cell(row=row, column=5).value
      quantity = sheetPiece.cell(row=row, column=4).value
      material = sheetPiece.cell(row=row, column=8).value
      unitWeight = f"{sheetPiece.cell(row=row, column=11).value:.1f}"
      totalWeight = f"{sheetPiece.cell(row=row, column=12).value:.1f}"

      pieceList.append(piece(cod, description, quantity, material, unitWeight, totalWeight))

    return pieceList
  
  # Lê informações do Peso das Peças e Soma dando o Peso total arrendonda sem casa decimal
  def getApproxTotalWeight(self):
    approxTotalWeight = 0.0
    
    for piece in self.pieces:
      approxTotalWeight = approxTotalWeight + float(piece.totalWeight)

    approxTotalWeight = str(round(approxTotalWeight))
    
    return approxTotalWeight
    
class materialList:
  def __init__(self, sheetMaterial):
    self.materials = self.readSheetMaterial(sheetMaterial)
    self.materialMeter = None
    self.materialSquareMeter = None
    self.getMaterialforMeter()

  # Lê informações da Planilha de Materiais
  def readSheetMaterial(self, sheetMaterial):
    materialList = []

    for row in range(2, sheetMaterial.max_row + 1):
      cell_value = sheetMaterial.cell(row=row, column=12).value
      if cell_value == None: break

      QTDE = sheetMaterial.cell(row=row, column=15).value
      QTDE = f"{QTDE:.1f}" if isinstance(QTDE, (int, float)) else QTDE
      
      position = sheetMaterial.cell(row=row, column=12).value
      description = sheetMaterial.cell(row=row, column=13).value
      unit = sheetMaterial.cell(row=row, column=14).value
      quantity = QTDE
      composition = sheetMaterial.cell(row=row, column=16).value
      totalWeight = f"{sheetMaterial.cell(row=row, column=17).value:.1f}"

      materialList.append(material(position, description, unit, quantity, composition, totalWeight))

    return materialList
  
  # Lê informações dos Materiais pelo Metro e Metro Quadrado para indicar o Material das Chapas e Material dos Perfil
  def getMaterialforMeter(self):
    materialMeter = set()
    materialSquareMeter = set()

    # Itere pela lista de dados e concatene o material com base na unidade
    for material in self.materials:
      if material.unit == 'm':
        materialMeter.add(material.composition)
      elif material.unit == 'm²':
        materialSquareMeter.add(material.composition)

    materialMeterJoin = " / ".join(materialMeter)
    materialSquareMeterJoin = " / ".join(materialSquareMeter)

    self.materialMeter = materialMeterJoin
    self.materialSquareMeter = materialSquareMeterJoin

class piece:
  def __init__(self, cod, description, quantity, material, unitWeight, totalWeight):
    self.cod = cod
    self.description = description
    self.quantity = quantity
    self.material = material
    self.unitWeight = unitWeight
    self.totalWeight = totalWeight

class material:
  def __init__(self, position, description, unit, quantity, composition, totalWeight):
    self.position = position
    self.description = description
    self.unit = unit
    self.quantity = quantity
    self.composition = composition
    self.totalWeight = totalWeight